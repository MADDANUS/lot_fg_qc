import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook('TEMPLATE lot fg qc OKE.xlsx', data_only=True)
ws = wb.active

print('=== COLUMN WIDTHS (cols 9-15) ===')
for c in range(9, 16):
    cl = get_column_letter(c)
    dim = ws.column_dimensions[cl]
    w = dim.width if dim.width else 10
    print(f'  Col {cl} (idx {c}): width={w}')

print()
print('=== ROW HEIGHTS (rows 1-18) ===')
for r in range(1, 19):
    h = ws.row_dimensions[r].height if ws.row_dimensions[r].height else 15
    print(f'  Row {r}: height={h}pt')

print()
print('=== MERGED CELLS affecting cols 9-15 ===')
for m in ws.merged_cells.ranges:
    if m.min_col >= 9 and m.max_col <= 15:
        vals = []
        for mr in range(m.min_row, m.max_row+1):
            for mc in range(m.min_col, m.max_col+1):
                v = ws[f'{get_column_letter(mc)}{mr}'].value
                if v: vals.append(str(v))
        val_str = ', '.join(vals) if vals else '(empty)'
        print(f'  {m} -> cols {m.min_col}-{m.max_col}, rows {m.min_row}-{m.max_row} = {val_str}')

print()
print('=== CELL BORDERS (right label) ===')
sep = '|'
for r in range(2, 19):
    for c in range(9, 16):
        cell = ws[f'{get_column_letter(c)}{r}']
        b = cell.border
        sides = []
        if b.top and b.top.style: sides.append('T')
        if b.bottom and b.bottom.style: sides.append('B')
        if b.left and b.left.style: sides.append('L')
        if b.right and b.right.style: sides.append('R')
        if sides:
            v = str(cell.value)[:15] if cell.value else '.'
            print(f'  {get_column_letter(c)}{r}: borders=[{sep.join(sides)}] val={v}')
