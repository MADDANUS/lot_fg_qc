"""
Analisis detail TEMPLATE lot kiri.xlsx dan TEMPLATE lot kanan.xlsx
Termasuk: ukuran kolom, tinggi baris, merged cells, border, nilai sel, background color.
"""
import openpyxl
from openpyxl.utils import get_column_letter

def analyze_sheet(filepath, label):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    print(f"\n{'='*70}")
    print(f"  {label}: {filepath}")
    print(f"  Sheet: '{ws.title}', Rows: {max_row}, Cols: {max_col}")
    print(f"{'='*70}")

    # Column widths
    print(f"\n--- Column Widths ---")
    total_w = 0
    for c in range(1, max_col + 1):
        cl = get_column_letter(c)
        dim = ws.column_dimensions.get(cl)
        w = dim.width if dim and dim.width else 8.43  # default Excel width
        total_w += w
        print(f"  Col {cl} ({c}): width = {w:.2f}")
    print(f"  TOTAL width units = {total_w:.2f}")

    # Row heights
    print(f"\n--- Row Heights ---")
    total_h = 0
    for r in range(1, max_row + 1):
        dim = ws.row_dimensions.get(r)
        h = dim.height if dim and dim.height else 15  # default
        total_h += h
        print(f"  Row {r}: height = {h:.2f} pt")
    print(f"  TOTAL height = {total_h:.2f} pt")

    # Merged cells
    print(f"\n--- Merged Cells ---")
    for m in sorted(ws.merged_cells.ranges, key=lambda x: (x.min_row, x.min_col)):
        start = f"{get_column_letter(m.min_col)}{m.min_row}"
        end = f"{get_column_letter(m.max_col)}{m.max_row}"
        val = ws[start].value
        val_str = str(val)[:40] if val else "(empty)"
        span_r = m.max_row - m.min_row + 1
        span_c = m.max_col - m.min_col + 1
        print(f"  {start}:{end}  (rowspan={span_r}, colspan={span_c})  = {val_str}")

    # Cell values, borders, fills
    print(f"\n--- Cell Details (row by row) ---")
    for r in range(1, max_row + 1):
        print(f"\n  Row {r}:")
        for c in range(1, max_col + 1):
            cell = ws[f"{get_column_letter(c)}{r}"]
            val = cell.value
            # Border info
            b = cell.border
            bsides = []
            if b.top and b.top.style: bsides.append(f"T({b.top.style})")
            if b.bottom and b.bottom.style: bsides.append(f"B({b.bottom.style})")
            if b.left and b.left.style: bsides.append(f"L({b.left.style})")
            if b.right and b.right.style: bsides.append(f"R({b.right.style})")
            # Fill info
            fill = cell.fill
            fill_str = ""
            if fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != '00000000':
                fill_str = f" fill={fill.fgColor.rgb}"
            # Font info
            font = cell.font
            font_parts = []
            if font.bold: font_parts.append("bold")
            if font.size: font_parts.append(f"size={font.size}")
            if font.name: font_parts.append(f"font={font.name}")
            font_str = f" [{', '.join(font_parts)}]" if font_parts else ""
            # Alignment
            align = cell.alignment
            align_parts = []
            if align.horizontal: align_parts.append(f"h={align.horizontal}")
            if align.vertical: align_parts.append(f"v={align.vertical}")
            if align.wrap_text: align_parts.append("wrap")
            if align.text_rotation: align_parts.append(f"rot={align.text_rotation}")
            align_str = f" align({', '.join(align_parts)})" if align_parts else ""

            if val is not None or bsides or fill_str:
                val_display = str(val)[:30] if val else "."
                border_str = f" border=[{' '.join(bsides)}]" if bsides else ""
                print(f"    {get_column_letter(c)}{r}: {val_display}{border_str}{fill_str}{font_str}{align_str}")

    wb.close()


# Analyze both files
analyze_sheet("TEMPLATE lot kiri.xlsx", "LABEL KIRI")
analyze_sheet("TEMPLATE lot kanan.xlsx", "LABEL KANAN")
