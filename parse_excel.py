import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('TEMPLATE lot fg qc OKE.xlsx', data_only=True)
ws = wb.active

html = '<table style="border-collapse: collapse; table-layout: fixed; width: 100%;">\n'
html += '  <colgroup>\n'
for c in range(1, ws.max_column + 1):
    col_letter = get_column_letter(c)
    dim = ws.column_dimensions[col_letter]
    width = dim.width if dim.width else 10
    html += f'    <col style="width: {width * 2}mm;">\n'
html += '  </colgroup>\n'

merged_cells = ws.merged_cells.ranges
skip_cells = set()

for r in range(1, ws.max_row + 1):
    height = ws.row_dimensions[r].height if ws.row_dimensions[r].height else 15
    html += f'  <tr style="height: {height}pt;">\n'
    for c in range(1, ws.max_column + 1):
        coord = f'{get_column_letter(c)}{r}'
        if coord in skip_cells:
            continue
            
        cell = ws[coord]
        colspan = 1
        rowspan = 1
        
        for merged in merged_cells:
            if coord == merged.start_cell.coordinate:
                colspan = merged.max_col - merged.min_col + 1
                rowspan = merged.max_row - merged.min_row + 1
                for mr in range(merged.min_row, merged.max_row + 1):
                    for mc in range(merged.min_col, merged.max_col + 1):
                        if not (mr == merged.min_row and mc == merged.min_col):
                            skip_cells.add(f'{get_column_letter(mc)}{mr}')
                break
                
        val = str(cell.value) if cell.value is not None else '&nbsp;'
        align = cell.alignment.horizontal if cell.alignment and cell.alignment.horizontal else 'left'
        valign = cell.alignment.vertical if cell.alignment and cell.alignment.vertical else 'top'
        bold = 'bold' if cell.font and cell.font.bold else 'normal'
        size = cell.font.size if cell.font else 10
        
        # approximate borders
        border_styles = []
        if cell.border:
            if cell.border.top.style: border_styles.append('border-top: 0.1mm solid #000;')
            if cell.border.bottom.style: border_styles.append('border-bottom: 0.1mm solid #000;')
            if cell.border.left.style: border_styles.append('border-left: 0.1mm solid #000;')
            if cell.border.right.style: border_styles.append('border-right: 0.1mm solid #000;')
            
        style = f'text-align: {align}; vertical-align: {valign}; font-weight: {bold}; font-size: {size}pt; padding: 1mm; {" ".join(border_styles)}'
        
        html += f'    <td colspan="{colspan}" rowspan="{rowspan}" style="{style}">{val}</td>\n'
    html += '  </tr>\n'
html += '</table>'

with open('excel_to_html.txt', 'w', encoding='utf-8') as f:
    f.write(html)
print('Done!')
