import openpyxl
from openpyxl.utils import get_column_letter
import json

wb = openpyxl.load_workbook('TEMPLATE lot fg qc OKE.xlsx', data_only=True)
ws = wb.active
merged_cells = ws.merged_cells.ranges

def gen_table(start_col, end_col, width_cm, height_cm):
    skip_cells = set()
    
    html = f'<table style="border-collapse: collapse; table-layout: fixed; width: {width_cm}cm; height: {height_cm}cm; overflow:hidden;">\n'
    html += '  <colgroup>\n'
    total_w = 0
    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        dim = ws.column_dimensions[col_letter]
        width = dim.width if dim.width else 10
        total_w += width
    
    for c in range(start_col, end_col + 1):
        col_letter = get_column_letter(c)
        dim = ws.column_dimensions[col_letter]
        width = dim.width if dim.width else 10
        pct = (width / total_w) * 100
        html += f'    <col style="width: {pct}%;">\n'
    html += '  </colgroup>\n'
    html += '  <tbody>\n'

    for r in range(1, 19): # Up to row 18
        html += f'  <tr>\n'
        for c in range(start_col, end_col + 1):
            coord = f'{get_column_letter(c)}{r}'
            if coord in skip_cells:
                continue
                
            cell = ws[coord]
            colspan = 1
            rowspan = 1
            
            for merged in merged_cells:
                if coord == merged.start_cell.coordinate:
                    colspan = min(merged.max_col, end_col) - merged.min_col + 1
                    rowspan = merged.max_row - merged.min_row + 1
                    for mr in range(merged.min_row, merged.max_row + 1):
                        for mc in range(merged.min_col, merged.max_col + 1):
                            if not (mr == merged.min_row and mc == merged.min_col):
                                skip_cells.add(f'{get_column_letter(mc)}{mr}')
                    break
                    
            val = str(cell.value) if cell.value is not None else '&nbsp;'
            if val == 'None': val = '&nbsp;'
            align = cell.alignment.horizontal if cell.alignment and cell.alignment.horizontal else 'left'
            valign = cell.alignment.vertical if cell.alignment and cell.alignment.vertical else 'top'
            bold = 'bold' if cell.font and cell.font.bold else 'normal'
            size = cell.font.size if cell.font else 10
            
            border_styles = []
            if cell.border:
                if cell.border.top.style: border_styles.append('border-top: 0.1mm solid #000;')
                if cell.border.bottom.style: border_styles.append('border-bottom: 0.1mm solid #000;')
                if cell.border.left.style: border_styles.append('border-left: 0.1mm solid #000;')
                if cell.border.right.style: border_styles.append('border-right: 0.1mm solid #000;')
                
            style = f'text-align: {align}; vertical-align: {valign}; font-weight: {bold}; font-size: {size}pt; padding: 1mm; {" ".join(border_styles)}'
            
            html += f'    <td colspan="{colspan}" rowspan="{rowspan}" style="{style}">{val}</td>\n'
        html += '  </tr>\n'
    html += '  </tbody>\n'
    html += '</table>\n'
    return html

left_html = gen_table(1, 7, 9.6, 7.2)
right_html = gen_table(9, 15, 9.5, 8.0)

# Replace placeholders
left_html = left_html.replace('015 - PT . -IJP/BS', '015 - PT. NIHON SEIKI INDONESIA - <?= esc($productName) ?>')
left_html = left_html.replace('>Lot Guarantee<', '><?php if ($lotGuarantee): ?>Lot<br>Guarantee<?php else: ?>&nbsp;<?php endif; ?><')
left_html = left_html.replace('>Lot SA<', '><?php if ($lotSa): ?>Lot SA<?php else: ?>&nbsp;<?php endif; ?><')
left_html = left_html.replace('>4M<', '><?php if ($is4m): ?>4M<?php else: ?>&nbsp;<?php endif; ?><')
left_html = left_html.replace('besi (partname)', '<?= esc($description) ?>')

# Left label: Lot No barcode (row 8 is label, row 9 is the barcode row - all cells are &nbsp;)
# Replace the Lot No label td to include lotno value, and inject barcode into the row below it
left_html = left_html.replace(
    '>Lot No.:<',
    '><b>Lot No.:</b> <?= esc($lotNoCombined) ?><'
)

# Row 9 under Lot No - inject barcode (6 colspan cells, all &nbsp; in row 9)
# Find the row after Lot No row (row index 9 in Excel = row after Lot No.:)
# The row with Qty: is row 9 in Excel - we inject barcode before it
left_html = left_html.replace(
    '>Qty:<',
    '><?= $barcodeSvg($lotNoCombined, $barcodeH) ?></td></tr><tr><td colspan="1" rowspan="2" style="text-align: left; vertical-align: top; font-size: 9.0pt; padding: 1mm; border-left: 0.1mm solid #000;"><b>Qty:</b> <?= esc((string)$lotQty) ?><'
)

# Ref No value (HDFKSHDFASHDFHI in row 12)
left_html = left_html.replace('HDFKSHDFASHDFHI', '<?= esc($refNo) ?><?= $barcodeSvg($refNo, $barcodeH * 0.8) ?>')

# Remarks value (OKE in row 13)
left_html = left_html.replace('>OKE<', '><?= esc($remark) ?><')

# Date/time (row 16)
left_html = left_html.replace('2026-02-09 16:05:00', '<?= $now ?>')

# QR Code left
left_html = left_html.replace(
    '<td colspan="2" rowspan="3" style="text-align: center; vertical-align: top; font-weight: normal; font-size: 11.0pt; padding: 1mm; border-top: 0.1mm solid #000; border-bottom: 0.1mm solid #000; border-left: 0.1mm solid #000; border-right: 0.1mm solid #000;">&nbsp;</td>',
    '<td colspan="2" rowspan="3" style="text-align: center; vertical-align: middle; padding: 1mm; border-top: 0.1mm solid #000; border-bottom: 0.1mm solid #000; border-left: 0.1mm solid #000; border-right: 0.1mm solid #000;"><?= $qrCodeImg($qrLeft, $qrSize) ?></td>'
)

right_html = right_html.replace('BACK NO F1814', 'BACK NO <?= esc($backNo) ?>')
right_html = right_html.replace('PART NO : 111111                                                                                                     ', '<div style="font-weight:bold;">PART NO : <span style="font-weight:normal;"><?= esc($itemCode) ?></span></div><div style="margin-top:1.5mm; text-align:left;"><?= $barcodeSvg($itemCode, $barcodeH * 0.7) ?></div>')
right_html = right_html.replace('LOTNO : 111111                                               ', '<div style="font-weight:bold;">LOTNO : <span style="font-weight:normal;"><?= esc($lotno) ?></span></div><div style="margin-top:1.5mm; text-align:left;"><?= $barcodeSvg($lotno, $barcodeH * 0.7) ?></div>')
right_html = right_html.replace('>757<', '><?= esc($operator) ?><')
right_html = right_html.replace('>YES<', '><?= esc($rohsFree ? \'YES\' : \'NO\') ?><')
right_html = right_html.replace('QTY : 500', '<div style="font-weight:bold;">QTY : <span style="font-weight:normal;"><?= esc((string) $lotQty) ?></span></div><div style="margin-top:1.5mm; text-align:left;"><?= $barcodeSvg((string) $lotQty, $barcodeH * 0.7) ?></div>')
right_html = right_html.replace('>WHFGL<', '><?= esc($warehouse) ?><')
right_html = right_html.replace('>OKE<', '><?= esc($remark) ?><')
right_html = right_html.replace('2026-09-02 00:00:00', '<?= $printDateLong ?>')
right_html = right_html.replace('HDFKSHDFASHDFHI', '<?= esc($refNo) ?>')
right_html = right_html.replace('HUHOEOSUFHSFHSUFHU1', '<div style="text-align:center;"><?= $barcodeSvg($refNo, $barcodeH * 0.6) ?></div>')
right_html = right_html.replace('2026-02-09 16:05:00', '<?= $now ?>')
right_html = right_html.replace('>78Y<', '><?= esc($shiftName ?: $userInitial) ?><')
right_html = right_html.replace(
    '<td colspan="4" rowspan="4" style="text-align: center; vertical-align: top; font-weight: normal; font-size: 11.0pt; padding: 1mm; border-top: 0.1mm solid #000; border-bottom: 0.1mm solid #000; border-left: 0.1mm solid #000; border-right: 0.1mm solid #000;">&nbsp;</td>',
    '<td colspan="4" rowspan="4" style="text-align: center; vertical-align: middle; padding: 1mm; border-top: 0.1mm solid #000; border-bottom: 0.1mm solid #000; border-left: 0.1mm solid #000; border-right: 0.1mm solid #000;"><div style="width:<?= $qrSize + 4 ?>px; margin: 0 auto;"><?= $qrCodeImg($qrRight, $qrSize) ?></div></td>'
)

with open('left_table.txt', 'w', encoding='utf-8') as f:
    f.write(left_html)
with open('right_table.txt', 'w', encoding='utf-8') as f:
    f.write(right_html)
print('Done!')
